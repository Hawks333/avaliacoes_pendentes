import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Configuração da página
st.set_page_config(page_title="Processamento de Planilhas", layout="wide")
st.title("📊 App Online de Processamento de Planilhas de Estudantes")

# Upload da planilha
uploaded_file = st.file_uploader("Escolha um arquivo Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    # Lê todas as abas
    xls = pd.ExcelFile(uploaded_file)
    st.write("Abas encontradas:", xls.sheet_names)
    
    # Seleção da aba
    sheet_name = st.selectbox("Selecione a aba para processar", xls.sheet_names)
    
    if sheet_name:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
        st.write("Pré-visualização da planilha:")
        st.dataframe(df.head())
        
        # Seleção da avaliativa
        avaliativa = st.selectbox("Selecione a Avaliativa", [1, 2, 3, 4])

        # Procurar todas as colunas relacionadas à avaliativa, ignorando tentativas
        colunas_avaliativa = []
        colunas_ignorar = []
        
        for col in df.columns:
            if f"avaliativa {avaliativa}" in col.lower():
                if "tentativas" in col.lower():
                    colunas_ignorar.append(col)
                else:
                    colunas_avaliativa.append(col)

        if colunas_avaliativa:
            st.write("Colunas encontradas para a avaliativa:", colunas_avaliativa)
            if colunas_ignorar:
                st.write("Colunas ignoradas (contêm 'Tentativas'):", colunas_ignorar)
            
            # Lista de colunas adicionais que queremos incluir
            colunas_adicionais = []
            colunas_padrao = ["DR", "Polo", "Nome"]
            
            # Verifica quais colunas adicionais existem no DataFrame
            for coluna in ["Etapa", "Sala", "Data último acesso"]:
                if coluna in df.columns:
                    colunas_adicionais.append(coluna)
                else:
                    st.warning(f"Coluna '{coluna}' não encontrada no DataFrame.")
            
            # Combina todas as colunas que vamos usar
            todas_colunas = colunas_padrao + colunas_adicionais + colunas_avaliativa
            
            # Filtra alunos que têm pelo menos um "--" em qualquer coluna da avaliativa (sem tentativas)
            mask = df[colunas_avaliativa].apply(lambda x: x.astype(str).str.contains("--")).any(axis=1)
            alunos_com_pendencia = df[mask][todas_colunas].copy()
            
            # Identifica as áreas com pendência para cada aluno
            def identificar_areas_pendentes(row):
                areas_pendentes = []
                for col in colunas_avaliativa:
                    if str(row[col]).strip() == "--":
                        # Extrai o nome da área (remove a parte da avaliativa)
                        area = col.replace(f"Avaliativa {avaliativa}", "").strip()
                        # Remove caracteres especiais no início
                        if area.startswith(('-', '–', '—', ':')):
                            area = area[1:].strip()
                        if area:  # Só adiciona se não for vazio
                            areas_pendentes.append(area)
                return ", ".join(areas_pendentes) if areas_pendentes else "Nenhuma"
            
            # Adiciona coluna com as áreas pendentes
            alunos_com_pendencia["Áreas com Pendência"] = alunos_com_pendencia.apply(identificar_areas_pendentes, axis=1)
            
            # Filtra apenas alunos que realmente têm pendências
            alunos_com_pendencia = alunos_com_pendencia[alunos_com_pendencia["Áreas com Pendência"] != "Nenhuma"]
            
            if not alunos_com_pendencia.empty:
                st.subheader("Estudantes com resultado pendente")
                
                # Mostra as colunas principais + áreas pendentes
                cols_to_show = colunas_padrao + colunas_adicionais + ["Áreas com Pendência"]
                df_to_show = alunos_com_pendencia[cols_to_show]
                
                # Aplicando cores alternadas
                def color_rows(row):
                    return ['background-color: #E0F7FA' if row.name % 2 == 0 else '']*len(row)
                
                st.dataframe(df_to_show.style.apply(color_rows, axis=1))
                
                # Opção de download
                towrite = BytesIO()
                
                # Criando Excel com cores
                with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
                    # Cria uma planilha resumida com informações principais
                    colunas_resumo = colunas_padrao + colunas_adicionais + ["Áreas com Pendência"]
                    resumo_download = alunos_com_pendencia[colunas_resumo].copy()
                    resumo_download.to_excel(writer, index=False, sheet_name="Pendências Resumidas")
                    
                    # Aplicando cores alternadas no Excel
                    ws = writer.sheets["Pendências Resumidas"]
                    fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
                    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
                        if idx % 2 == 0:
                            for cell in row:
                                cell.fill = fill
                    
                    # Cria uma segunda aba com detalhes completos
                    alunos_com_pendencia.to_excel(writer, index=False, sheet_name="Detalhes Completos")
                    
                    # Aplica formatação na aba de detalhes completos também
                    ws_detalhes = writer.sheets["Detalhes Completos"]
                    for idx, row in enumerate(ws_detalhes.iter_rows(min_row=2, max_row=ws_detalhes.max_row), start=0):
                        if idx % 2 == 0:
                            for cell in row:
                                cell.fill = fill
                
                towrite.seek(0)
                
                st.download_button(
                    label="⬇️ Baixar Relatório Completo",
                    data=towrite,
                    file_name=f"alunos_avaliativa_{avaliativa}_com_areas_pendentes.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Mostra estatísticas
                st.subheader("📈 Estatísticas das Pendências")
                total_pendencias = alunos_com_pendencia["Áreas com Pendência"].str.split(", ").explode().value_counts()
                st.write("Quantidade de pendências por área:")
                st.dataframe(total_pendencias)
                
                # Estatísticas adicionais por etapa, sala, etc.
                if "Etapa" in df.columns:
                    st.write("Pendências por Etapa:")
                    st.dataframe(alunos_com_pendencia["Etapa"].value_counts())
                
                if "Sala" in df.columns:
                    st.write("Pendências por Sala:")
                    st.dataframe(alunos_com_pendencia["Sala"].value_counts())
                
            else:
                st.info("Nenhum aluno com resultado pendente encontrado.")
        else:
            st.warning(f"Nenhuma coluna encontrada para a Avaliativa {avaliativa} (excluindo colunas de tentativas).")
